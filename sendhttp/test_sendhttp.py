import os
import pytest
import allure
import jsonpath
import openpyxl
import logging

from corehttp.corehttp import CoreHttp
from excel_reader.excelReader import ExcelReader1
from excel_reader.excelReader2 import ExcelReader2
from sql_check.sql_check import CheckSql
from mylogger.mylogger import MyLogger
from VAR.VAR import *


class Test_SendHttp:

    def setup_class(self):
        # 全局变量
        global ch, extract_dict, excel_path, excel, sheet, DB_conn, logger
        # 实例化api请求
        ch = CoreHttp()
        # 抽取字段存储
        extract_dict = {}
        # 文件路径变量
        excel_path = EXCEL_PATH
        # 操作excel
        excel = openpyxl.load_workbook(excel_path)
        sheet = excel['info']
        # 数据库初始化
        # DB_conn = CheckSql(user='dml_user', passwd='tiens_123', port=3306,
        #                    host="fat-bj-china-b2c-02.chfa9nngsipy.rds.cn-north-1.amazonaws.com.cn",
        #                    database="aishi_user_center")
        # 6、日志实例化
        my_logger = MyLogger()
        logger = my_logger.get_log("./mylogger/test.log", Clevel=logging.WARNING, Flevel=logging.DEBUG)

    # 文件读取接口数据
    models = ExcelReader1().excelReader()
    @pytest.mark.parametrize('model', models)
    # 传入的参数models是一个实例化对象列表，这个对象列表里的每个对象对应的是读到的excel的每一对象值
    def test_send_http(self, model):
        # 动态获取用例标题
        if model.case_desc:
            allure.dynamic.title(model.case_desc)
        # 动态获取业务模块feature
        if model.case_feature:
            allure.dynamic.feature(model.case_feature)
        # 动态获取业务功能模块story
        if model.case_story:
            allure.dynamic.story(model.case_story)
        # 动态获取备注
        if model.backup:
            allure.dynamic.description(model.backup)
        # 动态获取级别
        if model.case_level:
            allure.dynamic.severity(model.case_level)

        # 1、数据解析
        if model.num != 5:
            data_dict = {
                "url": model.url,
                "method": model.method,
                "headers": eval(model.headers),
                "data": eval(model.data),
                "para_type": model.para_type
            }
        else:
            data_dict = {
                "url": model.url,
                "method": model.method,
                "headers": eval(model.headers),
                "data": int(model.data),
                "para_type": model.para_type
            }

        # 2、接口请求+解包
        # print(f"执行第{model.num}条用例：【{model.case_desc}】")
        logger.info(f"执行第{model.num}条用例：【{model.case_desc}】")
        res = ch.core_http(**data_dict)
        # print(res.json())
        logger.debug(f"响应结果是：{res.json()}")

        # 3、数据抽取
        # 首先判断excel表中传入的抽取列表有数据
        if model.extract != "" and model.extract is not None:
            # 抽取参数可能为多个，所以需要循环抽取
            for ex in eval(model.extract):
                j_ex = jsonpath.jsonpath(res.json(), "$.." + ex)  # jsonpath得到的是一个list
                extract_dict.update({ex: j_ex[0]})
        # print(extract_dict)

        # 4、断言+返显
        r = model.num + 1
        print(jsonpath.jsonpath(res.json(), '$..' + model.assert_data))
        print(type(jsonpath.jsonpath(res.json(), '$..' + model.assert_data)))
        print(model.assert_value)
        print(type(model.assert_value))

        try:
            if model.assert_options == "包含":
                assert jsonpath.jsonpath(res.json(), '$..' + model.assert_data)
                print('包含：断言通过')

            elif model.assert_options == "大于":
                assert int(jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0]) > model.assert_value
                print('大于：断言通过')

            elif model.assert_options == "小于":
                assert int(res.json()[model.assert_data]) < model.assert_value
                print('小于：断言通过')

            elif model.assert_options == "等于":
                assert int(res.json()[model.assert_data]) == int(model.assert_value)
                print('等于：断言通过')

            else:
                print("断言条件设置异常")

        except Exception as e:
            print('断言失败/返回结果错误/校验字段错误')
            sheet.cell(r, 10).value = '断言失败/返回结果错误/校验字段错误'
            raise Exception("断言失败/返回结果错误/校验字段错误")

        else:
            sheet.cell(r, 10).value = '断言通过'

        finally:
            excel.save(excel_path)



        # 5、数据库校验
        # if model.sql is not None:
        #     try:
        #         if '{}' in model.sql:
        #             if model.sql_var is not None:
        #                 sql_select = model.sql.format(**eval(model.sql_var))
        #             else:
        #                 print("缺少sql语句查询变量，sql异常")
        #         else:
        #             sql_select = model.sql
        #         sql_res = DB_conn.chSql(sql_select)[0][0]
        #         if sql_res == str(model.sql_value):
        #             sheet.cell(r, 11).value = "通过"
        #         else:
        #             sheet.cell(r, 11).value = "失败"
        #     except Exception as e:
        #         print("sql有误/数据查询为空，请检查")
        #         sheet.cell(r, 11).value = 'sql有误/数据查询为空，请检查'
        #     finally:
        #         excel.save(excel_path)
        #         assert sql_res == str(model.sql_value), 'sql断言失败，case fail'
        #         print('数据库校验正确，测试用例执行成功')


# if __name__ == '__main__':
#     pytest.main([__file__,  # '-o', 'log_cli=true',   # 控制台输出日志，级别无法控制
#                  '--html=../htmlreport/report.html',
#                  '--alluredir', '../report', '--clean-alluredir'])
#     os.system('allure generate ../report -o ../allure_report/ --clean')
    # os.system('allure open ../allure_report/')

